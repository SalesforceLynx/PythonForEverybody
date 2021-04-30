from _collections import defaultdict
import openpyxl

SalesFunnel = defaultdict(list)

# class SalesFunnel(object):
#     def __init__(self, Index=None, LeadId=None, Email=None, LeadStatus=None, CreatedDate=None):
#         self.Index = Index
#         self.LeadId = LeadId
#         self.Email = Email
#         self.LeadStatus = LeadStatus
#         self.CreatedDate = CreatedDate

theFile = openpyxl.load_workbook('Report.xlsx')
allSheetNames = theFile.sheetnames

print("All sheet names {} " .format(theFile.sheetnames))

#myDict = {}

for sheet in allSheetNames:
    print("Current sheet name is {}" .format(sheet))
    currentSheet = theFile[sheet]


# openpyxl >= 2.4
header = [cell.value for cell in currentSheet[1]]

sfunnel = []

# for row in list(currentSheet.rows)[1:]:
#     for column in "ABCD":
#         values = {}
#         for key, cell in zip(header, row):
#             values[key] = cell.value
#         funnel = SalesFunnel(**values)
#         print(column)
        #print(funnel)
        #sfunnel.append(funnel)
        #print(sfunnel)


for row in range(1, currentSheet.max_row + 1):
    for column in "ADEF":  # Here you can add or reduce the columns
        cell_name = "{}{}".format(column, row)
        SalesFunnel[row] = [cell_name, currentSheet[cell_name].value]
        SalesFunnel[row].append(SalesFunnel[row])
    print(SalesFunnel)


# alternatively if the order is fixed

# for row in currentSheet.rows[1:]:
#     args = [cell.value for cell in row]
#     funnel = SalesFunnel(*args)
#     sfunnel.append(funnel)
#     print(sfunnel)