import openpyxl
import string


wb = openpyxl.load_workbook('Report.xlsx')
sheet = wb['Sheet1']

for column in range(1,sheet.max_column +1):
    headers.append(sheet.cell(column=column,row=1).value)

print(headers)

for row in range(2, sheet.max_row +1):
    line = dict()
for onehdr in selhdrs:
    cell_value = sheet.cell(column=headers.index(onehdr)+1,row=row).value
if type(cell_value).name == 'str':
    cell_value = cell_value.strip()
elif type(cell_value).name == 'int':
cell_value = str(cell_value)
elif cell_value is None:
cell_value = ''
line[onehdr] = cell_value
result_dict.append(line)

print(result_dict[0])