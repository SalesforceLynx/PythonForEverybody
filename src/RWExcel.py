import openpyxl

class SalesFunnel(object):
    def __init__(self, Index=None, LeadId=None, LeadStatus=None, CreatedDate=None):
        self.Index = Index
        self.LeadId = LeadId
        self.LeadStatus = LeadStatus
        self.CreatedDate = CreatedDate


theFile = openpyxl.load_workbook('Report.xlsx')
allSheetNames = theFile.sheetnames

print("All sheet names {} " .format(theFile.sheetnames))

myDict = {}

for sheet in allSheetNames:
    print("Current sheet name is {}" .format(sheet))
    currentSheet = theFile[sheet]


    for row in range(1, currentSheet.max_row + 1):
        for column in "ADEF":  # Here you can add or reduce the columns
            cell_name = "{}{}".format(column, row)
            myDict[row] = [cell_name, currentSheet[cell_name].value]
            print(cell_name)
            #print(myDict)
            #print("cell position {} has value {}".format(cell_name, currentSheet[cell_name].value))