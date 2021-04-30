import openpyxl
from _collections import defaultdict

person = defaultdict(dict)

theFile = openpyxl.load_workbook('Report.xlsx')
allSheetNames = theFile.sheetnames

for sheet in allSheetNames:
    print("Current sheet name is {}" .format(sheet))
    currentSheet = theFile[sheet]


#for row in currentSheet.iter_rows(min_row=2, max_col=5):
for row in range(1, currentSheet.max_row + 1):
    #i, li, em, ls, cd = (c.value for c in row)
    print(c.value for c in row)
    #if i not in person:
    #    person[i] = defaultdict(list)
    #person[i][li].append((em, ls, cd))

    #print(person)