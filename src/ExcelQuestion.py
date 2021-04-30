from _collections import defaultdict
import openpyxl

SalesFunnel = defaultdict(list)

theFile = openpyxl.load_workbook('Report.xlsx')
allSheetNames = theFile.sheetnames

print("All sheet names {} " .format(theFile.sheetnames))

for sheet in allSheetNames:
    print("Current sheet name is {}" .format(sheet))
    currentSheet = theFile[sheet]


sfunnel = []


for row in range(1, currentSheet.max_row + 1):
    for column in "ADEF":
        cell_name = "{}{}".format(column, row)
        SalesFunnel[row] = [cell_name, currentSheet[cell_name].value]
        SalesFunnel[row].append(SalesFunnel[row])
    print(SalesFunnel)

